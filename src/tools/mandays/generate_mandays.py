import json
import sys
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)

# ─── Colour / Style constants ─────────────────────────────────────────────────
BLUE_FILL  = PatternFill("solid", fgColor="ADD8E6")   # Light Blue (header)
GREEN_FILL = PatternFill("solid", fgColor="90EE90")   # Light Green (sprint / total)

THIN = Side(style="thin")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BOLD   = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ─── Style helpers ────────────────────────────────────────────────────────────

def apply_border(cell):
    cell.border = ALL_BORDER


def header_style(cell, is_role_col=False):
    cell.fill      = BLUE_FILL
    cell.font      = BOLD
    cell.border    = ALL_BORDER
    cell.alignment = CENTER


def write_row(ws, row_num, col_a_text, role_values: dict, roles,
              fill=None, bold=False, center_roles=True):
    """Write one formatted row; role_values maps role→value (omit = blank)."""
    c = ws.cell(row=row_num, column=1, value=col_a_text)
    c.font      = Font(bold=bold)
    c.alignment = LEFT
    c.border    = ALL_BORDER
    if fill:
        c.fill = fill

    for i, role in enumerate(roles, start=2):
        val = role_values.get(role, None)
        c = ws.cell(row=row_num, column=i, value=val)
        c.font      = Font(bold=bold)
        c.alignment = CENTER if center_roles else LEFT
        c.border    = ALL_BORDER
        if fill:
            c.fill = fill


# ─── Main generator ───────────────────────────────────────────────────────────

def generate_excel(data: dict, output_path: str) -> None:
    """Generate a Mandays Excel workbook from *data* and save to *output_path*."""
    project = data["project_info"]
    roles   = data["roles"]
    wbs     = data["work_breakdown_structure"]
    grand   = data["grand_total"]

    wb = Workbook()
    ws = wb.active
    ws.title = "WBS"

    # Column widths
    ws.column_dimensions["A"].width = 52
    for i, _ in enumerate(roles, start=2):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = 8

    row = 1

    # ─── Rows 1-3: Project Info ───────────────────────────────────────────────
    for label, value in [
        ("Nama Project", project["name"]),
        ("Start Date",   project["start_date"]),
        ("End Date",     project["end_date"]),
    ]:
        c = ws.cell(row=row, column=1, value=f"{label} : {value}")
        c.font = BOLD
        row += 1

    # ─── Row 4: Table Header ──────────────────────────────────────────────────
    c = ws.cell(row=row, column=1, value="WBS")
    header_style(c)
    for i, role in enumerate(roles, start=2):
        c = ws.cell(row=row, column=i, value=role)
        header_style(c, is_role_col=True)
    row += 1

    # ─── Sprint blocks ────────────────────────────────────────────────────────
    for sprint in wbs:
        sprint_label = f"{sprint['sprint_name']} ({sprint['period']})"
        write_row(ws, row, sprint_label, sprint.get("total_mandays", {}), roles,
                  fill=GREEN_FILL, bold=True)
        row += 1

        for feature in sprint["features"]:
            fg    = feature["feature_group"]
            tasks = feature.get("tasks", [])

            c = ws.cell(row=row, column=1, value=fg)
            c.font      = BOLD
            c.alignment = LEFT
            c.border    = ALL_BORDER
            for i, _ in enumerate(roles, start=2):
                c = ws.cell(row=row, column=i)
                c.border    = ALL_BORDER
                c.alignment = CENTER
            row += 1

            for task in tasks:
                write_row(ws, row, task["name"], task.get("mandays", {}), roles,
                          fill=None, bold=False)
                row += 1

    # ─── Grand Total ──────────────────────────────────────────────────────────
    write_row(ws, row, "Total", grand, roles, fill=GREEN_FILL, bold=True)

    wb.save(output_path)
    print(f"✅  Saved → {output_path}")


# ─── Standalone entry point ───────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python generate_mandays.py <input.json> <output.xlsx>")
        sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as f:
        data = json.load(f)
    generate_excel(data, sys.argv[2])
