"""
extract_wbs.py  –  WBS Excel → JSON converter
Usage: python extract_wbs.py <file.xlsx> [output.json]
"""

import json
import re
import sys
from pathlib import Path

import openpyxl


# ── Tiny helpers ───────────────────────────────────────────────────────────────

def cell_str(ws, row, col) -> str:
    """Return a cell value as a stripped string (empty string if None)."""
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def to_int(value) -> int | None:
    """Convert a cell value to a positive int; return None otherwise."""
    try:
        n = float(str(value).strip())
        return int(n) if n > 0 else None
    except (TypeError, ValueError):
        return None


def after_colon(text: str) -> str:
    """Extract the part after the first ':' and strip it."""
    return text.split(":", 1)[1].strip() if ":" in text else text.strip()


def sprint_parts(text: str) -> tuple[str, str]:
    """
    Split 'Sprint 2 (04- 17 Maret 2026)' into
    ('Sprint 2', '04- 17 Maret 2026').
    """
    m = re.match(r"^(Sprint\s*\d+)\s*\((.+)\)\s*$", text, re.IGNORECASE)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return text, ""


# ── Core logic ─────────────────────────────────────────────────────────────────

def read_excel(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # ── Project info (rows 1-3, format "Label : Value") ──
    project_info = {
        "name":       after_colon(cell_str(ws, 1, 1)),
        "start_date": after_colon(cell_str(ws, 2, 1)),
        "end_date":   after_colon(cell_str(ws, 3, 1)),
    }

    # ── Roles (row 4, columns B onwards) ──
    roles = []
    col = 2
    while True:
        r = cell_str(ws, 4, col)
        if not r:
            break
        roles.append(r)
        col += 1
    role_start_col = 2  # column index of the first role

    # ── Row helper: collect {role: value} for positive mandays ──
    def get_mandays(row):
        result = {}
        for i, role in enumerate(roles):
            n = to_int(ws.cell(row=row, column=role_start_col + i).value)
            if n is not None:
                result[role] = n
        return result

    # ── Iterate data rows (row 5 onwards) ──
    wbs = []
    grand_total = {}
    current_sprint = None
    current_feature = None

    for row in range(5, ws.max_row + 1):
        label = cell_str(ws, row, 1)

        if not label:                           # empty row → skip
            continue

        mandays = get_mandays(row)

        # Grand total
        if re.search(r"total", label, re.IGNORECASE):
            grand_total = mandays
            continue

        # Sprint header (name + period in the same cell)
        if re.search(r"sprint", label, re.IGNORECASE):
            if current_sprint:
                wbs.append(current_sprint)
            name, period = sprint_parts(label)
            current_sprint = {
                "sprint_name":   name,
                "period":        period,
                "total_mandays": mandays,
                "features":      [],
            }
            current_feature = None
            continue

        # Feature-group label (no mandays → category header)
        if not mandays:
            if current_sprint is not None:
                current_feature = {"feature_group": label, "tasks": []}
                current_sprint["features"].append(current_feature)
            continue

        # Task row (has mandays)
        task = {"name": label, "mandays": mandays}

        if current_sprint is None:
            # Task before any sprint → individual task
            wbs.append({"task_name": label, "type": "individual_task", "mandays": mandays})
        elif current_feature is not None:
            current_feature["tasks"].append(task)
        else:
            # Task inside a sprint but no feature group yet
            current_sprint.setdefault("ungrouped_tasks", []).append(task)

    # Commit last sprint
    if current_sprint:
        wbs.append(current_sprint)

    return {
        "project_info":             project_info,
        "roles":                    roles,
        "work_breakdown_structure": wbs,
        "grand_total":              grand_total,
    }


def save_json(data: dict, path: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Saved → {Path(path).resolve()}")


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_wbs.py <file.xlsx> [output.json]")
        sys.exit(1)

    excel_path  = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else "output.json"

    if not Path(excel_path).exists():
        print(f"File not found: {excel_path}")
        sys.exit(1)

    data = read_excel(excel_path)
    save_json(data, output_path)

    # Brief summary
    sprints = sum(1 for x in data["work_breakdown_structure"] if "sprint_name" in x)
    print(f"Project : {data['project_info']['name']}")
    print(f"Roles   : {', '.join(data['roles'])}")
    print(f"Sprints : {sprints}")
    print(f"Total MD: {sum(data['grand_total'].values())}")


if __name__ == "__main__":
    main()
