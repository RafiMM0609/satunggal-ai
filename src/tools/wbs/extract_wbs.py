#!/usr/bin/env python3
import argparse
import json
import logging
from datetime import datetime, timedelta
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

def _get_color_hex(cell):
    """Return color hex (uppercase) for a cell's fill start_color if available."""
    try:
        sc = cell.fill.start_color
    except Exception:
        return None
    if sc is None:
        return None
    # prefer rgb
    rgb = getattr(sc, "rgb", None)
    if rgb:
        return str(rgb).upper()
    # some openpyxl versions use .index
    idx = getattr(sc, "index", None)
    if idx:
        return str(idx).upper()
    # fallback to theme or tint
    th = getattr(sc, "theme", None)
    if th is not None:
        return str(th).upper()
    return None


def parse_excel_to_json(file_path, wbs_start_row=6, date_header_row=5, date_start_col=2, date_end_col=9):
    """Parse workbook into WBS JSON using color rules:

    - Category: cell A fill = #92D050 (FF92D050)
    - Sub-Category: cell A fill = #B7DEE8 (FFB7DEE8)
    - Task rows: regular cells under a sub-category
    - For each Task, check columns B..I and if cell fill = #31869B (FF31869B)
      then include the date from the header row (`date_header_row`) into `active_days`.

    Returns list of categories with sub_categories and tasks.
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    data_wbs = []
    current_cat = None
    current_sub = None

    # Normalize hex check helper
    def is_hex_equal(col_hex, target_hex):
        if not col_hex:
            return False
        s = col_hex.replace("#", "").upper()
        # some values may be like 'FF92D050' or '92D050'
        return s.endswith(target_hex.replace("#", "").upper())

    # iterate rows starting at the WBS area
    for row in ws.iter_rows(min_row=wbs_start_row, max_col=date_end_col):
        cell_a = row[0]
        cell_a_val = cell_a.value
        fill_hex = _get_color_hex(cell_a) or ""

        # 1. Category (green #92D050)
        if is_hex_equal(fill_hex, "FF92D050") or is_hex_equal(fill_hex, "92D050"):
            current_cat = {"category": str(cell_a_val).strip() if cell_a_val else "", "sub_categories": []}
            data_wbs.append(current_cat)
            current_sub = None
            continue

        # 2. Sub-Category (light blue #B7DEE8)
        if is_hex_equal(fill_hex, "FFB7DEE8") or is_hex_equal(fill_hex, "B7DEE8"):
            current_sub = {"name": str(cell_a_val).strip() if cell_a_val else "", "tasks": []}
            if current_cat is None:
                # create uncategorized container if missing
                current_cat = {"category": "Uncategorized", "sub_categories": []}
                data_wbs.append(current_cat)
            current_cat["sub_categories"].append(current_sub)
            continue

        # 3. Task row: has value in column A and is not a colored category/sub row
        if cell_a_val and str(cell_a_val).strip():
            task_name = str(cell_a_val).strip()
            task = {"task_name": task_name, "active_days": []}

            # check date columns B..I (row[1] .. row[date_end_col-1])
            # row is a tuple of cells length = date_end_col
            for idx, cell in enumerate(row[date_start_col - 1:date_end_col - 0]):
                # compute actual column index in worksheet
                col_idx = date_start_col + idx
                color = _get_color_hex(cell) or ""
                if is_hex_equal(color, "FF31869B") or is_hex_equal(color, "31869B"):
                    # get date value from header row
                    date_val = ws.cell(row=date_header_row, column=col_idx).value
                    # coerce to integer day when possible
                    if isinstance(date_val, datetime):
                        day = date_val.day
                    else:
                        try:
                            s = str(date_val).strip()
                            if s.isdigit():
                                day = int(s)
                            else:
                                # if value like '20' with whitespace or other, try to extract digits
                                import re

                                m = re.search(r"(\d{1,2})", s)
                                day = int(m.group(1)) if m else s
                        except Exception:
                            day = date_val
                    task["active_days"].append(day)

            # attach task to current_sub (preferred) or create default sub
            if current_sub is None:
                # create default sub if needed
                if current_cat is None:
                    current_cat = {"category": "Uncategorized", "sub_categories": []}
                    data_wbs.append(current_cat)
                # ensure there is a 'General' sub
                # find or create 'General'
                gen = None
                for s in current_cat["sub_categories"]:
                    if s.get("name") == "General":
                        gen = s
                        break
                if gen is None:
                    gen = {"name": "General", "tasks": []}
                    current_cat["sub_categories"].append(gen)
                gen["tasks"].append(task)
            else:
                current_sub["tasks"].append(task)

    return data_wbs

ID_MONTHS = [
    "Januari",
    "Februari",
    "Maret",
    "April",
    "Mei",
    "Juni",
    "Juli",
    "Agustus",
    "September",
    "Oktober",
    "November",
    "Desember",
]


ID_WEEKDAY_ID = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]


def get_sprint_days(start_date, working_days=10):
    """Return a list of workday objects starting from `start_date`.

    - `start_date` may be a `datetime`, `date` or ISO string `YYYY-MM-DD`.
    - Includes `start_date` if it falls on a weekday (Mon-Fri).
    - Skips Saturdays and Sundays.
    - If the sequence passes end-of-month it continues naturally into next month.

    Returns list of dicts: {"tanggal": int, "bulan": "Februari", "hari": "Senin"}
    """
    # coerce start_date
    if start_date is None:
        raise ValueError("start_date is required")
    if isinstance(start_date, str):
        try:
            sd = datetime.fromisoformat(start_date).date()
        except Exception:
            # try to parse common '18 Februari 2026' formats
            s = start_date.strip()
            for i, m in enumerate(ID_MONTHS, start=1):
                if m.lower() in s.lower():
                    parts = s.replace(m, str(i)).split()
                    nums = [p for p in parts if p.isdigit()]
                    if len(nums) >= 2:
                        day = int(nums[0])
                        year = int(nums[-1])
                        sd = datetime(year, i, day).date()
                        break
            else:
                raise
    elif isinstance(start_date, datetime):
        sd = start_date.date()
    else:
        sd = start_date

    res = []
    cur = sd
    # count weekdays (Mon-Fri)
    while len(res) < int(working_days):
        if cur.weekday() < 5:  # 0..4 => Mon..Fri
            res.append({
                "tanggal": cur.day,
                "bulan": ID_MONTHS[cur.month - 1],
                "hari": ID_WEEKDAY_ID[cur.weekday()],
            })
        cur = cur + timedelta(days=1)
    return res


def is_colored(cell):
    if cell is None:
        return False
    fill = getattr(cell, "fill", None)
    if not fill:
        return False
    # patternType can be 'solid' for colored cells
    pattern = getattr(fill, "patternType", None)
    if not pattern:
        return False
    fg = getattr(fill, "fgColor", None)
    if not fg:
        return False
    # Try multiple indicators: rgb, indexed, type
    rgb = getattr(fg, "rgb", None)
    if rgb:
        s = str(rgb).upper()
        # common 'white' values in openpyxl
        if s in ("00000000", "FFFFFFFF", "00FFFFFF", "FFFFFF", "00000000"):
            return False
        return True
    # if no rgb present, but theme/indexed set, consider it colored
    idx = getattr(fg, "indexed", None)
    if idx is not None and idx != 64:
        return True
    return False


def find_timeline_header(sheet, search_start_col=4, look_rows=40):
    max_col = sheet.max_column
    # find the row that contains many day numbers (1..31)
    for r in range(1, min(look_rows, sheet.max_row) + 1):
        candidate = []
        for c in range(search_start_col, max_col + 1):
            cell = sheet.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            # accept int, datetime, or numeric string
            if isinstance(v, int) and 1 <= v <= 31:
                candidate.append((c, int(v)))
            else:
                try:
                    if isinstance(v, datetime):
                        candidate.append((c, int(v.day)))
                    else:
                        vs = str(v).strip()
                        if vs.isdigit() and 1 <= int(vs) <= 31:
                            candidate.append((c, int(vs)))
                except Exception:
                    continue
        if len(candidate) >= 3:
            cols = [it[0] for it in candidate]
            days = [it[1] for it in candidate]

            # For each date column, try to find a month label in rows above that column
            col_month = {}
            for i, c in enumerate(cols):
                month_text = None
                # scan upwards to find a month name
                for rr in range(max(1, r - 6), 0, -1):
                    txt = sheet.cell(row=rr, column=c).value
                    if txt and isinstance(txt, str):
                        s = txt.strip()
                        for m in ID_MONTHS:
                            if m.lower() in s.lower():
                                month_text = m
                                break
                        if month_text:
                            break
                col_month[c] = month_text

            # look for sprint label between a few rows above and the days row
            sprint_name = None
            for rr in range(max(1, r - 4), r):
                for c in range(search_start_col, max_col + 1):
                    txt = sheet.cell(row=rr, column=c).value
                    if txt and isinstance(txt, str) and 'sprint' in txt.lower():
                        nxt = None
                        # try to find a textual sprint name nearby
                        for cc in range(c, min(max_col, c + 6) + 1):
                            v = sheet.cell(row=rr, column=cc).value
                            if v and isinstance(v, str) and v.strip() != '':
                                nxt = v
                                break
                        sprint_name = nxt if nxt else txt
                        break
                if sprint_name:
                    break

                return r, cols, days, col_month, sprint_name
    return None, [], [], {}, None


def extract_project_info(sheet):
    info = {"project_name": None, "start_date": None, "end_date": None}
    seen_start = 0
    for r in range(1, 10):
        a = sheet.cell(row=r, column=1).value
        if not a:
            continue
        s = str(a).strip()
        # handle "Label : value" in same cell
        if ":" in s:
            left, right = s.split(":", 1)
            key = left.strip().lower()
            val = right.strip()
        else:
            key = s.lower()
            val = sheet.cell(row=r, column=2).value
            if val is None:
                val = ""
        if "nama project" in key or "project" in key:
            if val:
                info["project_name"] = str(val).strip()
        if "start date" in key:
            if seen_start == 0:
                info["start_date"] = _coerce_date(val)
                seen_start += 1
            else:
                info["end_date"] = _coerce_date(val)
    return info


def _coerce_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # try parse common formats like '18 Februari 2026' or '20 April 2026'
    try:
        # replace Indonesian month names with month numbers
        for i, m in enumerate(ID_MONTHS, start=1):
            if m.lower() in s.lower():
                # extract day and year
                parts = s.replace(m, str(i)).split()
                nums = [p for p in parts if p.isdigit()]
                if len(nums) >= 2:
                    day = int(nums[0])
                    year = int(nums[-1])
                    return datetime(year, i, day).strftime("%Y-%m-%d")
        # fallback try generic parse
        try:
            dt = datetime.fromisoformat(s)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return s
    except Exception:
        return s


def parse_sheet(path, sheet_name=None):
    wb = load_workbook(path, data_only=True)
    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb[wb.sheetnames[0]]

    proj = extract_project_info(sheet)

    header_row, cols, days, col_month, sprint_name = find_timeline_header(sheet)
    if header_row is None:
        raise SystemExit("Tidak menemukan header timeline. Pastikan kolom tanggal ada pada baris awal.")

    # map col index to day
    col_to_day = {c: days[i] for i, c in enumerate(cols)}

    # derive month_text from per-column detection if available
    month_text = None
    # col_month mapping returned by find_timeline_header may contain per-column month hints
    # prefer the most common month among detected columns
    detected_months = [m for m in col_month.values() if m]
    if detected_months:
        # choose most frequent
        from collections import Counter

        cnt = Counter(detected_months)
        month_text = cnt.most_common(1)[0][0]
    else:
        # fallback: try find a month label anywhere above the header_row
        for rr in range(1, header_row):
            for cc in range(1, sheet.max_column + 1):
                txt = sheet.cell(row=rr, column=cc).value
                if txt and isinstance(txt, str):
                    s = txt.strip()
                    for m in ID_MONTHS:
                        if m.lower() in s.lower():
                            month_text = m
                            break
                    if month_text:
                        break
            if month_text:
                break

    # attempt to detect sprint_name more robustly: scan a few rows above the header
    if not sprint_name:
        import re

        for rr in range(max(1, header_row - 6), header_row):
            for cc in range(1, sheet.max_column + 1):
                txt = sheet.cell(row=rr, column=cc).value
                if not txt or not isinstance(txt, str):
                    continue
                s = txt.strip()
                if re.search(r"\bsprint\b", s, re.I):
                    # prefer full nearby label if empty
                    # if cell contains 'Sprint' only, check right neighbor
                    if s.lower() in ("sprint", "sprint:"):
                        neigh = sheet.cell(row=rr, column=min(sheet.max_column, cc + 1)).value
                        if neigh and isinstance(neigh, str) and neigh.strip():
                            sprint_name = neigh.strip()
                        else:
                            sprint_name = s
                    else:
                        sprint_name = s
                    break
            if sprint_name:
                break

    # build chronological timeline entries from detected cols/days
    day_vals = [col_to_day[c] for c in cols]

    # determine starting month/year for inference
    if month_text:
        try:
            base_month = ID_MONTHS.index(month_text) + 1
        except Exception:
            base_month = None
    else:
        base_month = None

    base_year = None
    try:
        if proj and proj.get("start_date"):
            sd = proj.get("start_date")
            if isinstance(sd, str) and '-' in sd:
                base_year = int(sd.split("-")[0])
            elif isinstance(sd, datetime):
                base_year = sd.year
    except Exception:
        base_year = None

    if base_month is None:
        # fallback to project start month or current month
        if proj and proj.get("start_date"):
            try:
                if isinstance(proj.get("start_date"), str):
                    base_month = int(proj.get("start_date").split("-")[1])
                else:
                    base_month = proj.get("start_date").month
            except Exception:
                base_month = datetime.today().month
        else:
            base_month = datetime.today().month

    if base_year is None:
        if proj and proj.get("start_date") and isinstance(proj.get("start_date"), str) and '-' in proj.get("start_date"):
            try:
                base_year = int(proj.get("start_date").split("-")[0])
            except Exception:
                base_year = datetime.today().year
        else:
            base_year = datetime.today().year

    timeline = []
    cur_month = base_month
    cur_year = base_year
    prev_day = None
    for d in day_vals:
        if prev_day is None:
            # first element keep base_month
            pass
        else:
            # if day value decreases, assume month boundary crossed
            try:
                if int(d) < int(prev_day):
                    cur_month += 1
                    if cur_month > 12:
                        cur_month = 1
                        cur_year += 1
            except Exception as exc:
                logger.debug("extract_wbs: month-boundary detection failed for day=%r: %s", d, exc)
        is_weekday = True
        try:
            dt = datetime(cur_year, cur_month, int(d))
            is_weekday = dt.weekday() < 5
        except Exception:
            # if invalid (e.g. 31 in shorter month), try next month
            try:
                tmp_month = cur_month + 1
                tmp_year = cur_year
                if tmp_month > 12:
                    tmp_month = 1
                    tmp_year += 1
                dt = datetime(tmp_year, tmp_month, int(d))
                is_weekday = dt.weekday() < 5
                # adjust current month if we consumed next month
                cur_month = tmp_month
                cur_year = tmp_year
            except Exception:
                is_weekday = True

        timeline.append({
            "date": int(d),
            "month": ID_MONTHS[cur_month - 1],
            "year": cur_year,
            "is_weekday": bool(is_weekday),
            # sprint will be filled after we detect per-column sprint labels
            "sprint": None,
        })
        prev_day = d

    # build month spans (consecutive ranges in timeline) to help Excel merging later
    month_spans = []
    if timeline:
        cur_m = timeline[0]["month"]
        start_idx = 0
        for i, item in enumerate(timeline[1:], start=1):
            if item["month"] != cur_m:
                month_spans.append({"month": cur_m, "start": start_idx, "end": i - 1})
                cur_m = item["month"]
                start_idx = i
        month_spans.append({"month": cur_m, "start": start_idx, "end": len(timeline) - 1})

    # Detect per-column sprint labels by scanning rows above header_row
    sprint_labels = []  # list of tuples (col_index, text)
    import re
    for rr in range(max(1, header_row - 6), header_row):
        for cc in range(1, sheet.max_column + 1):
            txt = sheet.cell(row=rr, column=cc).value
            if txt and isinstance(txt, str) and re.search(r"\bsprint\b", txt, re.I):
                sprint_labels.append((cc, txt.strip()))

    # deduplicate by column, keep leftmost/latest occurrence
    sprint_labels_sorted = sorted({c: t for c, t in sprint_labels}.items()) if sprint_labels else []

    # assign sprint label to each timeline item based on nearest left sprint label
    sprint_per_index = [None] * len(timeline)
    if sprint_labels_sorted:
        # convert to list of (col, text)
        sprint_positions = [(c, t) for c, t in sprint_labels_sorted]
        # for each timeline column (cols list), find last sprint pos with col <= that col
        for i, col in enumerate(cols):
            label = None
            for cpos, text in sprint_positions:
                if cpos <= col:
                    label = text
                else:
                    break
            sprint_per_index[i] = label

    # fill sprint field in timeline items
    for i, item in enumerate(timeline):
        item["sprint"] = sprint_per_index[i] if i < len(sprint_per_index) else None

    # build sprint spans from sprint_per_index
    sprint_spans = []
    if any(sprint_per_index):
        cur_s = sprint_per_index[0]
        start_idx = 0
        for i, s in enumerate(sprint_per_index[1:], start=1):
            if s != cur_s:
                sprint_spans.append({"sprint": cur_s, "start": start_idx, "end": i - 1})
                cur_s = s
                start_idx = i
        sprint_spans.append({"sprint": cur_s, "start": start_idx, "end": len(sprint_per_index) - 1})
    else:
        sprint_spans = []

    # Build sprints array using sprint_spans; each sprint contains its days list
    sprints = []
    if sprint_spans:
        for idx, sp in enumerate(sprint_spans, start=1):
            name = sp.get("sprint") or (sprint_name or f"Sprint {idx}")
            days = []
            for item in timeline[sp["start"] : sp["end"] + 1]:
                days.append({
                    "date": item["date"],
                    "month": item["month"],
                    "year": item.get("year"),
                    "is_weekday": item.get("is_weekday", True),
                })
            sprints.append({"sprint_name": name, "days": days})
    else:
        # fallback: single sprint containing all timeline days
        name = sprint_name or "Sprint 1"
        days = [
            {"date": it["date"], "month": it["month"], "year": it.get("year"), "is_weekday": it.get("is_weekday", True)}
            for it in timeline
        ]
        sprints.append({"sprint_name": name, "days": days})

    timeline_config = {
        "sprints": sprints
    }

    # helper to read color properties
    def _color_props(cell):
        sc = getattr(cell.fill, 'start_color', None)
        if not sc:
            return {"rgb": None, "index": None, "theme": None}
        return {"rgb": getattr(sc, 'rgb', None), "index": getattr(sc, 'index', None), "theme": getattr(sc, 'theme', None)}

    def is_category_green(cell):
        p = _color_props(cell)
        if p['rgb']:
            if str(p['rgb']).upper().endswith('92D050') or str(p['rgb']).upper().endswith('FF92D050'):
                return True
        if p['index']:
            # some files use rgb in index field
            try:
                if str(p['index']) == 'FF92D050' or str(p['index']).endswith('92D050'):
                    return True
            except Exception as exc:
                logger.debug("extract_wbs: is_category_green color-index check failed: %s", exc)
        p = _color_props(cell)
        if p['rgb']:
            if str(p['rgb']).upper().endswith('B7DEE8') or str(p['rgb']).upper().endswith('FFB7DEE8'):
                return True
        if p['index']:
            # in this workbook light-blue appears as index==7
            try:
                if int(p['index']) == 7:
                    return True
            except Exception:
                if str(p['index']).upper().endswith('B7DEE8'):
                    return True
        return False

    def is_active_blue(cell):
        p = _color_props(cell)
        if p['rgb']:
            ru = str(p['rgb']).upper()
            if ru.endswith('31869B') or ru.endswith('FF31869B') or ru.endswith('00B0F0') or ru.endswith('FF00B0F0'):
                return True
        if p['index']:
            try:
                if int(p['index']) == 7:
                    return True
            except Exception:
                iu = str(p['index']).upper()
                if iu.endswith('31869B') or iu.endswith('FF31869B') or iu.endswith('00B0F0') or iu.endswith('FF00B0F0'):
                    return True
        # also check fgColor in case start_color was empty
        fg = getattr(cell.fill, 'fgColor', None)
        if fg:
            fr = getattr(fg, 'rgb', None)
            fi = getattr(fg, 'index', None)
            if fr and (str(fr).upper().endswith('31869B') or str(fr).upper().endswith('00B0F0')):
                return True
            try:
                if fi and int(fi) == 7:
                    return True
            except Exception:
                if fi and (str(fi).upper().endswith('31869B') or str(fi).upper().endswith('00B0F0')):
                    return True
        return False

    # Layout: column A contains Category / Sub-Category / Task names
    category_col = 1
    task_col = 1

    max_row = sheet.max_row
    data = {}
    current_category = None
    current_sub = None
    # We will iterate rows below the header_row to collect categories/subcategories/tasks
    for r in range(header_row + 1, max_row + 1):
        cat_cell = sheet.cell(row=r, column=category_col)
        cat_val = cat_cell.value
        # detect category header if category cell is green
        if cat_val and is_category_green(cat_cell):
            current_category = str(cat_val).strip()
            data.setdefault(current_category, {"tasks": [], "sub": {}})
            current_sub = None
            continue

        # detect subcategory: often in column A but light-blue filled; try column A first
        sub_cell = sheet.cell(row=r, column=category_col)
        sub_val = sub_cell.value
        if sub_val and is_sub_lightblue(sub_cell) and current_category:
            current_sub = str(sub_val).strip()
            data.setdefault(current_category, {"tasks": [], "sub": {}})
            data[current_category]["sub"].setdefault(current_sub, [])
            continue

        # tasks are in column A for this layout
        task_val = cat_val
        if not task_val or (is_category_green(cat_cell) or is_sub_lightblue(cat_cell)):
            continue
        task_name = str(task_val).strip()
        active_days = []
        for c in cols:
            cell = sheet.cell(row=r, column=c)
            if is_active_blue(cell):
                active_days.append(col_to_day[c])

        if current_category is None:
            current_category = "Uncategorized"
            data.setdefault(current_category, {"tasks": [], "sub": {}})

        if current_sub:
            data[current_category]["sub"].setdefault(current_sub, []).append({"task_name": task_name, "active_days": active_days})
        else:
            data[current_category]["tasks"].append({"task_name": task_name, "active_days": active_days})

    wbs = []
    for cat, content in data.items():
        entry = {"category": cat, "is_header": True}
        if content.get("sub"):
            subcats = []
            for subname, tasks in content["sub"].items():
                subcats.append({"name": subname, "tasks": tasks})
            # include any tasks without subcategory as a default subgroup
            if content["tasks"]:
                subcats.insert(0, {"name": "General", "tasks": content["tasks"]})
            entry["sub_categories"] = subcats
        else:
            entry["tasks"] = content["tasks"]
        wbs.append(entry)

    result = {
        "project_info": proj,
        "timeline_config": timeline_config,
        "wbs_data": wbs,
    }

    return result


def main():
    p = argparse.ArgumentParser()
    p.add_argument("input", help="path to excel file (.xlsx)")
    p.add_argument("-s", "--sheet", help="sheet name (optional)")
    p.add_argument("-o", "--output", help="output json file", default=None)
    args = p.parse_args()

    out = parse_sheet(args.input, args.sheet)
    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, indent=2)
        print("Saved to", args.output)
    else:
        print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
