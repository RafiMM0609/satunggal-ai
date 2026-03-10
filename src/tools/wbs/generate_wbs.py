"""
json_to_excel.py
Generates a WBS Excel file from the JSON output produced by excel_to_json.py.
Usage:
    python json_to_excel.py output.json wbs_output.xlsx
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour constants
# ---------------------------------------------------------------------------
COLOR_HEADER_GREEN  = "FF92D050"   # Category header row
COLOR_SUBCAT_BLUE   = "FFB7DEE8"   # Sub-category row
COLOR_ACTIVE_CELL   = "FF31869B"   # Timeline cell that is active
COLOR_HEADER_BG     = "FFD9D9D9"   # Month / sprint header background
COLOR_BORDER        = "FF000000"


# ---------------------------------------------------------------------------
# Helper – thin border
# ---------------------------------------------------------------------------
def thin_border():
    s = Side(style="thin", color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


# ---------------------------------------------------------------------------
# Core: find which column indices (0-based inside flat_days) should be active
# for a given task, using contiguous-subsequence matching.
#
# Strategy:
#   1. Build flat_days = [(date, month, year), ...] for all timeline days.
#   2. For each *group* of tasks (category or sub-category), pick the task
#      with the most active_days to anchor the sprint window via contiguous
#      matching.  All other tasks in the group then search only within that
#      window (±small buffer), which handles single-date tasks correctly.
# ---------------------------------------------------------------------------

def find_contiguous_match(dates: list, flat_days: list,
                           search_start: int = 0,
                           search_end: int | None = None) -> list[int]:
    """
    Return a list of 0-based indices into flat_days where the sequence of
    date-numbers in `dates` appears as a contiguous subsequence.
    Returns [] if no match is found.
    """
    if not dates:
        return []

    if search_end is None:
        search_end = len(flat_days) - 1

    n = len(flat_days)
    m = len(dates)

    for i in range(search_start, search_end - m + 2):
        if i < 0 or i >= n:
            continue
        if flat_days[i]["date"] == dates[0]:
            matched = True
            for j in range(1, m):
                ni = i + j
                if ni > search_end or ni >= n or flat_days[ni]["date"] != dates[j]:
                    matched = False
                    break
            if matched:
                return list(range(i, i + m))
    return []


def determine_window(tasks: list, flat_days: list) -> tuple[int, int]:
    """
    Determine the sprint window (start_idx, end_idx) for a group of tasks
    by using the task with the largest active_days list (most information).
    Adds a small buffer around the found window.
    Buffer is 10 days – generous enough to always capture the full sprint.
    """
    best_task = None
    for t in tasks:
        ad = t.get("active_days", [])
        if len(ad) > 1:
            if best_task is None or len(ad) > len(best_task.get("active_days", [])):
                best_task = t

    if best_task:
        cols = find_contiguous_match(best_task["active_days"], flat_days)
        if cols:
            buf = 15
            return (max(0, min(cols) - buf), min(len(flat_days) - 1, max(cols) + buf))

    # Fallback: full range
    return (0, len(flat_days) - 1)


def active_columns_for_group(tasks: list, flat_days: list) -> dict[str, list[int]]:
    """
    Returns a dict mapping task_name -> list of 0-based column indices in flat_days.
    """
    result = {}
    window_start, window_end = determine_window(tasks, flat_days)

    for task in tasks:
        name = task["task_name"]
        ad   = task.get("active_days", [])
        if not ad:
            result[name] = []
            continue

        cols = find_contiguous_match(ad, flat_days, window_start, window_end)

        # Fallback to full search if windowed search failed
        if not cols:
            cols = find_contiguous_match(ad, flat_days)

        result[name] = cols

    return result


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def style_cell(cell, *, value=None, bold=False, fill_color=None,
               align_center=False, font_color="FF000000", font_size=10,
               border=True, wrap=False):
    if value is not None:
        cell.value = value
    cell.font = Font(bold=bold, color=font_color, size=font_size)
    if fill_color:
        cell.fill = make_fill(fill_color)
    cell.alignment = Alignment(
        horizontal="center" if align_center else "left",
        vertical="center",
        wrap_text=wrap,
    )
    if border:
        cell.border = thin_border()


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_excel(data: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "WBS"

    project = data["project_info"]
    sprints  = data["timeline_config"]["sprints"]
    wbs_data = data["wbs_data"]

    # -----------------------------------------------------------------------
    # Build flat list of all days across all sprints
    # flat_days[i] = {"date": int, "month": str, "year": int, "sprint": str}
    # -----------------------------------------------------------------------
    flat_days: list[dict] = []
    for sprint in sprints:
        for day in sprint["days"]:
            flat_days.append({
                "date":   day["date"],
                "month":  day["month"],
                "year":   day["year"],
                "sprint": sprint["sprint_name"],
            })

    total_cols = len(flat_days)
    # Column A = 1 (task name), timeline starts at column 4 (D)
    # so that excel_to_json.py's find_timeline_header (search_start_col=4) can detect it
    DATA_COL_OFFSET = 4   # openpyxl column index for 1st timeline column

    # -----------------------------------------------------------------------
    # ROW 1-3  –  Project metadata
    # -----------------------------------------------------------------------
    ws["A1"].value = f"Nama Project : {project['project_name']}"
    ws["A2"].value = f"Start Date   : {project['start_date']}"
    # The parser (extract_project_info) matches 'start date' twice via a counter;
    # the 2nd match is stored as end_date. So we keep the label 'Start Date' here.
    ws["A3"].value = f"Start Date   : {project['end_date']}"  # end_date value
    for r in range(1, 4):
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)

    # -----------------------------------------------------------------------
    # ROW 4  –  Month headers (merged per month)
    # -----------------------------------------------------------------------
    MONTH_ROW  = 4
    SPRINT_ROW = 5
    DATE_ROW   = 6
    WBS_START  = 7

    # Group columns by month
    month_groups: list[tuple[int, int, str]] = []   # (start_col, end_col, month_label)
    cur_month = None
    cur_start = None

    for i, day in enumerate(flat_days):
        col = i + DATA_COL_OFFSET
        month_label = f"{day['month']} {day['year']}"
        if month_label != cur_month:
            if cur_month is not None:
                month_groups.append((cur_start, col - 1, cur_month))
            cur_month  = month_label
            cur_start  = col
    month_groups.append((cur_start, DATA_COL_OFFSET + total_cols - 1, cur_month))

    for start_col, end_col, label in month_groups:
        if start_col == end_col:
            cell = ws.cell(row=MONTH_ROW, column=start_col, value=label)
        else:
            ws.merge_cells(
                start_row=MONTH_ROW, start_column=start_col,
                end_row=MONTH_ROW,   end_column=end_col,
            )
            cell = ws.cell(row=MONTH_ROW, column=start_col, value=label)
        style_cell(cell, bold=True, fill_color="FFD9D9D9", align_center=True)

    # -----------------------------------------------------------------------
    # ROW 5  –  Sprint headers (merged per sprint)
    # -----------------------------------------------------------------------
    sprint_groups: list[tuple[int, int, str]] = []
    cur_sprint = None
    cur_start  = None

    for i, day in enumerate(flat_days):
        col = i + DATA_COL_OFFSET
        if day["sprint"] != cur_sprint:
            if cur_sprint is not None:
                sprint_groups.append((cur_start, col - 1, cur_sprint))
            cur_sprint = day["sprint"]
            cur_start  = col
    sprint_groups.append((cur_start, DATA_COL_OFFSET + total_cols - 1, cur_sprint))

    for start_col, end_col, label in sprint_groups:
        if start_col == end_col:
            cell = ws.cell(row=SPRINT_ROW, column=start_col, value=label)
        else:
            ws.merge_cells(
                start_row=SPRINT_ROW, start_column=start_col,
                end_row=SPRINT_ROW,   end_column=end_col,
            )
            cell = ws.cell(row=SPRINT_ROW, column=start_col, value=label)
        style_cell(cell, bold=True, fill_color="FFBDD7EE", align_center=True)

    # -----------------------------------------------------------------------
    # ROW 6  –  Date numbers
    # -----------------------------------------------------------------------
    for i, day in enumerate(flat_days):
        col  = i + DATA_COL_OFFSET
        cell = ws.cell(row=DATE_ROW, column=col, value=day["date"])
        style_cell(cell, bold=True, fill_color="FFDCE6F1", align_center=True)

    # -----------------------------------------------------------------------
    # Column A header
    # -----------------------------------------------------------------------
    ws.cell(row=MONTH_ROW,  column=1, value="WBS")
    ws.cell(row=SPRINT_ROW, column=1, value="")
    ws.cell(row=DATE_ROW,   column=1, value="Task")
    for r in (MONTH_ROW, SPRINT_ROW, DATE_ROW):
        style_cell(ws.cell(row=r, column=1), bold=True, fill_color="FFD9D9D9",
                   align_center=True)

    # -----------------------------------------------------------------------
    # WBS rows
    # -----------------------------------------------------------------------
    current_row = WBS_START

    def fill_timeline_row(row: int, active_col_indices: list[int]):
        """Paint timeline cells; leave inactive cells white with border."""
        for i in range(total_cols):
            col  = i + DATA_COL_OFFSET
            cell = ws.cell(row=row, column=col)
            if i in active_col_indices:
                style_cell(cell, fill_color=COLOR_ACTIVE_CELL, align_center=True)
            else:
                style_cell(cell)   # border only, no fill

    def write_category_header(name: str):
        nonlocal current_row
        cell = ws.cell(row=current_row, column=1, value=name)
        style_cell(cell, bold=True, fill_color=COLOR_HEADER_GREEN)
        fill_timeline_row(current_row, [])
        # Paint the header row across all timeline cols with the same green
        for i in range(total_cols):
            col  = i + DATA_COL_OFFSET
            c    = ws.cell(row=current_row, column=col)
            style_cell(c, fill_color=COLOR_HEADER_GREEN)
        current_row += 1

    def write_subcat_header(name: str):
        nonlocal current_row
        cell = ws.cell(row=current_row, column=1, value=name)
        style_cell(cell, bold=True, fill_color=COLOR_SUBCAT_BLUE)
        for i in range(total_cols):
            col  = i + DATA_COL_OFFSET
            c    = ws.cell(row=current_row, column=col)
            style_cell(c, fill_color=COLOR_SUBCAT_BLUE)
        current_row += 1

    def write_task_row(task_name: str, active_col_indices: list[int]):
        nonlocal current_row
        cell = ws.cell(row=current_row, column=1, value=task_name)
        style_cell(cell, wrap=True)
        fill_timeline_row(current_row, active_col_indices)
        current_row += 1

    # Iterate wbs_data
    for category in wbs_data:
        cat_name  = category["category"]
        is_header = category.get("is_header", False)

        if is_header:
            write_category_header(cat_name)

        # ---- Tasks directly under category (no sub_categories) ----
        if "tasks" in category:
            tasks = category["tasks"]
            col_map = active_columns_for_group(tasks, flat_days)
            for task in tasks:
                write_task_row(task["task_name"], col_map[task["task_name"]])

        # ---- Sub-categories ----
        if "sub_categories" in category:
            for sub in category["sub_categories"]:
                write_subcat_header(sub["name"])
                tasks   = sub["tasks"]
                col_map = active_columns_for_group(tasks, flat_days)
                for task in tasks:
                    write_task_row(task["task_name"], col_map[task["task_name"]])

    # -----------------------------------------------------------------------
    # Column widths
    # -----------------------------------------------------------------------
    ws.column_dimensions["A"].width = 45
    for i in range(total_cols):
        col_letter = get_column_letter(i + DATA_COL_OFFSET)
        ws.column_dimensions[col_letter].width = 4

    # Row heights
    for r in range(MONTH_ROW, current_row):
        ws.row_dimensions[r].height = 20
    # Task rows – taller to accommodate wrapping
    for r in range(WBS_START, current_row):
        ws.row_dimensions[r].height = 30

    # -----------------------------------------------------------------------
    # Freeze panes at B7 so task names and headers stay visible
    # -----------------------------------------------------------------------
    ws.freeze_panes = "B7"

    wb.save(output_path)
    print(f"Excel file saved: {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python json_to_excel.py <input.json> <output.xlsx>")
        sys.exit(1)

    json_path  = sys.argv[1]
    excel_path = sys.argv[2]

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    generate_excel(data, excel_path)
